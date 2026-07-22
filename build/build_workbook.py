#!/usr/bin/env python3
"""
Twin Visit Logger — reference workbook builder (Phase 2).

Single source of truth for the upgraded data structure: column order, dropdown
lists, Google-Sheets-compatible formulas, the Cherry Opportunity Board, the
Exception Queue, the Migration Log, and the migrated pilot + test records.

Output: build/Twin_Visit_Logger_DEV_reference.xlsx

NOTES
- This produces a REFERENCE workbook. The live upgrade is applied to the Google
  Sheets dev copy by the Apps Script setup() routine (apps-script/Setup.gs),
  which mirrors the schema defined here.
- Formulas here use LibreOffice/Excel-portable syntax so they can be validated
  locally with scripts/recalc.py:
    * TEXTJOIN is written as _xlfn.TEXTJOIN (required by openpyxl for xlsx).
    * Ranges are bounded ($X$2:$X$1000), not open-ended ($X$2:$X).
  In Google Sheets the equivalent formulas use plain TEXTJOIN and may use
  open-ended ranges; see docs/Data-Dictionary.md for the Google variants.
- No Excel-only / Google-only exclusive functions are used in the Data sheet
  (only IF, AND, OR, TODAY, MAX, COUNTIFS, NETWORKDAYS, SUBSTITUTE, TRIM, LOWER,
  TEXTJOIN) so the same logic works on both platforms.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import datetime

LAST = 1000  # bounded-range last row for portable COUNTIFS

# ---------------------------------------------------------------------------
# 1. DROPDOWN VOCABULARIES  (exact values from the approved spec; assumptions flagged)
# ---------------------------------------------------------------------------
DROPDOWNS = {
    "Visit Status":        ["Scheduled", "Completed", "Canceled", "Reschedule Needed"],
    "Current Stage":       ["Visit Scheduled", "Visit Completed — Needs Review",
                             "Offer Preparation", "Offer Sent", "Active Negotiation",
                             "Verbal Agreement", "Contract Sent", "Contract Signed",
                             "Long-Term Nurture", "Lost / Closed Out"],
    "Assigned Owner":      ["Jonathan", "Kyle", "Cherry", "Juan", "JM"],
    "Assigned Visitor":    ["Juan", "Kyle", "Cherry", "Jonathan", "JM",
                            "Cesar", "Jose Herrera", "Manny Morales", "Lily", "Alan Hernandez"],  # owners + legacy field reps (assumption)
    "Gift Approval Owner": ["Cherry", "Juan"],
    "Updated By":          ["Jonathan", "Kyle", "Cherry", "Juan", "JM", "Apps Script", "Import"],
    "Final Disposition":   ["Contracted", "Lost", "Long-Term Nurture", "Closed Out"],
    "Gift Status":         ["Not Reviewed", "Recommended", "Approved", "Sent", "Not Appropriate"],
    "Blocker":             ["Price", "Title", "Tenant", "Family", "Access", "Timing",
                            "Documents", "Property Condition", "Seller Unresponsive", "Other"],
    "Lead Source":         ["Direct Mail", "Direct Mail - Postcard", "PPC", "TV", "Facebook",
                            "SEO", "PPL - Property Leads", "PPL - Motivated Leads"],
    # --- assumption-based lists (documented in Data-Dictionary.md) ---
    "Offer Status":        ["Not Started", "In Preparation", "Sent", "Countered",
                            "Accepted", "Rejected", "Withdrawn"],
    "Occupancy Status":    ["Owner-Occupied", "Tenant-Occupied", "Vacant", "Unknown"],
    "Property Condition":  ["Excellent", "Good", "Fair", "Poor", "Distressed"],
    "Seller Timeline":     ["ASAP", "30 days", "60 days", "90+ days", "Unknown"],
    "Offer Received Confirmation": ["Yes", "No"],
    "Transaction Handoff Status":  ["Not Ready", "Ready for Handoff", "Handed Off to JM", "JM Confirmed"],
    "REI Update Required":  ["Yes", "No"],
    "REI Update Completed": ["Yes", "No"],
    "Source":               ["Manual", "Apps Script", "Import", "TEST"],
    "Data Quality Status":  ["OK", "Incomplete", "Exception"],
    "Stalled Status":       ["Yes", "No"],
}

# ---------------------------------------------------------------------------
# 2. COLUMN SCHEMA  (group, header, dropdown_key or None, formula_template or None)
#    formula tokens:  {Header} -> $<col><row>   ;  [[Header]] -> $<col>$2:$<col>$LAST
# ---------------------------------------------------------------------------
COLUMNS = [
    # PROPERTY INFORMATION
    ("Property",   "Property ID",            None, None),
    ("Property",   "Property Address",       None, None),
    ("Property",   "Normalized Address",     None,
        '=IF({Property Address}="","",TRIM(LOWER('
        'SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({Property Address},",","")'
        ',".","")," apt "," "),"#","")," unit "," "))))'),
    ("Property",   "Seller Name",            None, None),
    ("Property",   "Phone",                  None, None),
    ("Property",   "Email",                  None, None),
    ("Property",   "Lead Source",            "Lead Source", None),
    ("Property",   "REI BlackBook Link",     None, None),
    # VISIT INFORMATION
    ("Visit",      "Visit Date",             None, None),
    ("Visit",      "Visit Time",             None, None),
    ("Visit",      "Visit Status",           "Visit Status", None),
    ("Visit",      "Assigned Visitor",       "Assigned Visitor", None),
    ("Visit",      "Visit Notes",            None, None),
    ("Visit",      "Property Condition",     "Property Condition", None),
    ("Visit",      "Occupancy Status",       "Occupancy Status", None),
    ("Visit",      "Photos Link",            None, None),
    ("Visit",      "Video Link",             None, None),
    ("Visit",      "File Link",              None, None),
    # SELLER INFORMATION
    ("Seller",     "Seller Motivation",      None, None),
    ("Seller",     "Seller Timeline",        "Seller Timeline", None),
    ("Seller",     "Asking Price",           None, None),
    ("Seller",     "Price Expectation",      None, None),
    ("Seller",     "Seller Concerns",        None, None),
    # OFFER INFORMATION
    ("Offer",      "Approved Offer Amount",  None, None),
    ("Offer",      "Offer Status",           "Offer Status", None),
    ("Offer",      "Offer Prepared Date",    None, None),
    ("Offer",      "Offer Sent Date",        None, None),
    ("Offer",      "Offer Received Confirmation", "Offer Received Confirmation", None),
    ("Offer",      "Counteroffer Amount",    None, None),
    # FOLLOW-UP INFORMATION
    ("Follow-up",  "Last Contact Date",      None, None),
    ("Follow-up",  "Last Contact Result",    None, None),
    ("Follow-up",  "Next Action",            None, None),
    ("Follow-up",  "Next Action Due Date",   None, None),
    ("Follow-up",  "Assigned Owner",         "Assigned Owner", None),
    ("Follow-up",  "Blocker",                "Blocker", None),
    ("Follow-up",  "Days Since Last Activity", None,
        '=IF({Property Address}="","",IF(AND({Last Contact Date}="",{Last Updated Date}="",{Visit Date}=""),"",'
        'TODAY()-MAX({Last Contact Date},{Last Updated Date},{Visit Date})))'),
    ("Follow-up",  "Days Overdue",           None,
        '=IF({Property Address}="","",IF({Next Action Due Date}="","",'
        'IF(TODAY()>{Next Action Due Date},TODAY()-{Next Action Due Date},0)))'),
    ("Follow-up",  "Stalled Status",         None,
        '=IF({Property Address}="","",IF(OR({Current Stage}="Lost / Closed Out",{Current Stage}="Long-Term Nurture",'
        '{Current Stage}="Contract Signed"),"No",'
        'IF(MAX({Last Contact Date},{Last Updated Date},{Visit Date})=0,"No",'
        'IF(NETWORKDAYS(MAX({Last Contact Date},{Last Updated Date},{Visit Date}),TODAY())-1>=3,"Yes","No"))))'),
    # RELATIONSHIP INFORMATION
    ("Relationship","Gift Status",           "Gift Status", None),
    ("Relationship","Gift Recommendation Reason", None, None),
    ("Relationship","Gift Approval Owner",   "Gift Approval Owner", None),
    ("Relationship","Gift Sent Date",        None, None),
    # CLOSEOUT INFORMATION
    ("Closeout",   "Current Stage",          "Current Stage", None),
    ("Closeout",   "Final Disposition",      "Final Disposition", None),
    ("Closeout",   "Closeout Reason",        None, None),
    ("Closeout",   "Contract Sent Date",     None, None),
    ("Closeout",   "Contract Signed Date",   None, None),
    ("Closeout",   "Transaction Handoff Status", "Transaction Handoff Status", None),
    # COMPUTED FLAGS
    ("Computed",   "Missing Required Fields", None,
        '=IF(OR({Property Address}="",{Current Stage}="Lost / Closed Out"),"",_xlfn.TEXTJOIN(", ",TRUE,'
        'IF({Property Address}="","Property Address",""),'
        'IF({Current Stage}="","Current Stage",""),'
        'IF({Next Action}="","Next Action",""),'
        'IF({Next Action Due Date}="","Next Action Due Date",""),'
        'IF({Assigned Owner}="","Assigned Owner",""),'
        'IF({REI BlackBook Link}="","REI BlackBook Link","")))'),
    ("Computed",   "Duplicate Address Flag", None,
        '=IF({Normalized Address}="","",IF(COUNTIFS([[Normalized Address]],{Normalized Address},'
        '[[Current Stage]],"<>Lost / Closed Out")>1,"Duplicate",""))'),
    ("Computed",   "Opportunity Priority",   None,
        '=IF({Property Address}="","",'
        'IF({Current Stage}="Verbal Agreement",100,'
        'IF({Current Stage}="Contract Sent",95,'
        'IF({Current Stage}="Active Negotiation",85,'
        'IF({Current Stage}="Offer Sent",70,'
        'IF({Current Stage}="Offer Preparation",60,'
        'IF({Current Stage}="Visit Completed — Needs Review",50,'
        'IF({Current Stage}="Visit Scheduled",30,'
        'IF({Current Stage}="Long-Term Nurture",10,'
        'IF({Current Stage}="Contract Signed",5,0)))))))))'
        '+IF({Days Overdue}="",0,MIN({Days Overdue},20))'
        '+IF({Stalled Status}="Yes",5,0))'),
    # SYSTEM INFORMATION
    ("System",     "Created Date",           None, None),
    ("System",     "Last Updated Date",      None, None),
    ("System",     "Updated By",             "Updated By", None),
    ("System",     "Source",                 "Source", None),
    ("System",     "Data Quality Status",    None,
        '=IF({Property Address}="","",'
        'IF({Exception Reason}<>"","Exception",'
        'IF({Missing Required Fields}<>"","Incomplete","OK")))'),
    ("System",     "Exception Reason",       None,
        '=IF({Property Address}="","",_xlfn.TEXTJOIN(" | ",TRUE,'
        # rule 1 & 2: completed visit
        'IF(AND({Visit Status}="Completed",{Visit Notes}=""),"Completed visit missing Visit Notes",""),'
        'IF(AND({Visit Status}="Completed",{Seller Motivation}=""),"Completed visit missing Seller Motivation (or add Exception note)",""),'
        # rule 3: offer sent
        'IF(AND({Current Stage}="Offer Sent",OR({Approved Offer Amount}="",{Offer Sent Date}="")),"Offer Sent needs Approved Offer Amount + Offer Sent Date",""),'
        # rule 4: active negotiation
        'IF(AND({Current Stage}="Active Negotiation",OR({Last Contact Result}="",{Next Action}="",{Assigned Owner}="",{Next Action Due Date}="")),"Active Negotiation needs Last Contact Result + Next Action + Owner + Due Date",""),'
        # rule 5: contract sent
        'IF(AND({Current Stage}="Contract Sent",{Contract Sent Date}="",{File Link}=""),"Contract Sent needs Contract Sent Date or File Link",""),'
        # rule 6: contract signed
        'IF(AND({Current Stage}="Contract Signed",{Contract Signed Date}=""),"Contract Signed needs Contract Signed Date",""),'
        # rule 7: nurture future date
        'IF(AND({Current Stage}="Long-Term Nurture",OR({Next Action Due Date}="",{Next Action Due Date}<=TODAY())),"Long-Term Nurture needs an exact FUTURE follow-up date",""),'
        # rule 8: lost/closed out
        'IF(AND({Current Stage}="Lost / Closed Out",OR({Final Disposition}="",{Closeout Reason}="")),"Lost / Closed Out needs Final Disposition + Closeout Reason",""),'
        # rule 9: gift sent requires approval
        'IF(AND({Gift Status}="Sent",{Gift Approval Owner}=""),"Gift marked Sent without recorded approval",""),'
        # rule 10: duplicate
        'IF({Duplicate Address Flag}="Duplicate","Duplicate active record for this address","")))'),
    ("System",     "REI Update Required",    "REI Update Required", None),
    ("System",     "REI Update Completed",   "REI Update Completed", None),
]

HEADERS = [c[1] for c in COLUMNS]
COL_IDX = {h: i + 1 for i, h in enumerate(HEADERS)}   # header -> 1-based col index
COL_LET = {h: get_column_letter(i + 1) for i, h in enumerate(HEADERS)}

GROUP_FILL = {
    "Property":     "D9E1F2",
    "Visit":        "E2EFDA",
    "Seller":       "FFF2CC",
    "Offer":        "FCE4D6",
    "Follow-up":    "DDEBF7",
    "Relationship": "EAD1DC",
    "Closeout":     "D9D2E9",
    "Computed":     "F2F2F2",
    "System":       "E7E6E6",
}


def render_formula(template, row):
    """Substitute {Header} -> $<col><row> and [[Header]] -> $<col>$2:$<col>$LAST."""
    out = template
    # ranges first
    for h in HEADERS:
        out = out.replace(f"[[{h}]]", f"${COL_LET[h]}$2:${COL_LET[h]}${LAST}")
    for h in HEADERS:
        out = out.replace(f"{{{h}}}", f"${COL_LET[h]}{row}")
    return out


# ---------------------------------------------------------------------------
# 3. PILOT MIGRATION  (10 most-recent legacy records) + TEST records
#    Values are migrated conservatively. Uncertain mappings -> exception queue.
# ---------------------------------------------------------------------------
def D(y, m, d):
    return datetime.datetime(y, m, d)

# Each record is a dict header->value. Only set what maps confidently.
PILOT = [
    # row2 Cyn Ku — scheduled visit; no REI link -> will flag Incomplete
    {"Property ID": "TVL-0001", "Property Address": "2607 Gimelli Pl, Apt 115, San Jose, CA 95133",
     "Seller Name": "Cyn Ku", "Phone": "(510) 284-7867", "Lead Source": "PPL - Property Leads",
     "Visit Date": D(2026,7,24), "Visit Status": "Scheduled", "Assigned Visitor": "Juan",
     "Visit Notes": "Cyn still interested in selling; 2nd property she wants to sell to us.",
     "Current Stage": "Visit Scheduled", "Next Action": "Conduct scheduled visit & log outcome",
     "Next Action Due Date": D(2026,7,24), "Assigned Owner": "Juan",
     "Created Date": D(2026,7,24), "Last Updated Date": D(2026,7,24), "Updated By": "Import",
     "Source": "Import", "REI Update Required": "Yes"},
    # row3 Steve Giorgi — scheduled
    {"Property ID": "TVL-0002", "Property Address": "1253 Edgewood Rd, Redwood City, CA 94062",
     "Seller Name": "Steve Giorgi", "Phone": "(650) 333-8189", "Lead Source": "PPL - Property Leads",
     "Visit Date": D(2026,7,22), "Visit Status": "Scheduled", "Assigned Visitor": "Juan",
     "Visit Notes": "Appointment booked; note outcome after visit.",
     "Current Stage": "Visit Scheduled", "Next Action": "Conduct scheduled visit & log outcome",
     "Next Action Due Date": D(2026,7,22), "Assigned Owner": "Juan",
     "Created Date": D(2026,7,22), "Last Updated Date": D(2026,7,22), "Updated By": "Import",
     "Source": "Import", "REI Update Required": "Yes"},
    # row4 Carmen Green — offer sent, missing amount/date -> EXCEPTION (rule 3)
    {"Property ID": "TVL-0003", "Property Address": "519 S 17th St, Richmond, CA 94804",
     "Seller Name": "Carmen Green", "Phone": "(916) 752-5759", "Lead Source": "PPL - Motivated Leads",
     "Visit Date": D(2026,7,20), "Visit Status": "Completed", "Assigned Visitor": "Juan",
     "Visit Notes": "Offer sent via email + SignNow contract; Juan following up directly.",
     "Seller Motivation": "Interested in selling (motivation detail not captured in legacy)",
     "Current Stage": "Offer Sent", "Offer Status": "Sent",
     "Next Action": "Follow up on sent offer", "Next Action Due Date": D(2026,7,23),
     "Assigned Owner": "Juan", "Last Contact Date": D(2026,7,20),
     "Last Contact Result": "Offer emailed; awaiting seller response",
     "Created Date": D(2026,7,20), "Last Updated Date": D(2026,7,20), "Updated By": "Import",
     "Source": "Import", "REI Update Required": "Yes"},
    # row5 Dorol Conrad — lost
    {"Property ID": "TVL-0004", "Property Address": "5 Lancaster Cir Apt 121, Bay Point, CA 94565",
     "Seller Name": "Dorol Conrad", "Phone": "(415) 370-9841", "Lead Source": "PPL - Property Leads",
     "Visit Date": D(2026,7,20), "Visit Status": "Canceled",
     "Visit Notes": "Equity 100% but property already listed on the MLS.",
     "Current Stage": "Lost / Closed Out", "Final Disposition": "Lost",
     "Closeout Reason": "We're Passing — already listed on MLS",
     "Created Date": D(2026,7,20), "Last Updated Date": D(2026,7,20), "Updated By": "Import", "Source": "Import"},
    # row6 Jon Box — lost
    {"Property ID": "TVL-0005", "Property Address": "15340 Canyon 2 Rd, Guerneville, CA 95446",
     "Seller Name": "Jon Box", "Phone": "(707) 481-7040", "Lead Source": "PPL - Property Leads",
     "Visit Date": D(2026,7,20), "Visit Status": "Canceled", "Assigned Visitor": "Juan",
     "Visit Notes": "Passing on this lead.",
     "Current Stage": "Lost / Closed Out", "Final Disposition": "Lost",
     "Closeout Reason": "We're Passing",
     "Created Date": D(2026,7,18), "Last Updated Date": D(2026,7,18), "Updated By": "Import", "Source": "Import"},
    # row7 Chris Giro — lost, completed
    {"Property ID": "TVL-0006", "Property Address": "18 Hampton Rd, Occidental, CA 95465",
     "Seller Name": "Chris J. Giro", "Phone": "(707) 292-9001", "Lead Source": "PPC",
     "Visit Date": D(2026,7,18), "Visit Status": "Completed", "Assigned Visitor": "Juan",
     "Visit Notes": "House is structurally failing; passing.",
     "Seller Motivation": "Motivated but property condition too poor for our criteria",
     "Current Stage": "Lost / Closed Out", "Final Disposition": "Lost",
     "Closeout Reason": "We're Passing — structural condition",
     "Created Date": D(2026,7,16), "Last Updated Date": D(2026,7,18), "Updated By": "Import", "Source": "Import"},
    # row8 Yvette Rose — lost, completed
    {"Property ID": "TVL-0007", "Property Address": "4107 Randolph Ave, Oakland, CA 94602",
     "Seller Name": "Yvette D Rose", "Phone": "(510) 457-6727", "Lead Source": "TV",
     "Visit Date": D(2026,7,18), "Visit Status": "Completed",
     "Visit Notes": "Heavily tenant-occupied; City of Oakland code violations.",
     "Seller Motivation": "Good location but tenant/code issues; passing",
     "Occupancy Status": "Tenant-Occupied", "Blocker": "Tenant",
     "Current Stage": "Lost / Closed Out", "Final Disposition": "Lost",
     "Closeout Reason": "We're Passing — tenant-occupied + code violations",
     "Created Date": D(2026,7,16), "Last Updated Date": D(2026,7,17), "Updated By": "Import", "Source": "Import"},
    # row9 Liam — lost
    {"Property ID": "TVL-0008", "Property Address": "16125 Bittner Rd, Occidental, CA 95465",
     "Seller Name": "Liam", "Phone": "(530) 545-1943", "Lead Source": "PPL - Property Leads",
     "Visit Date": D(2026,7,15), "Visit Status": "Canceled",
     "Visit Notes": "Internal cancellation of confirmed visit; passing.",
     "Current Stage": "Lost / Closed Out", "Final Disposition": "Lost",
     "Closeout Reason": "We're Passing — internal cancellation",
     "Created Date": D(2026,7,15), "Last Updated Date": D(2026,7,15), "Updated By": "Import", "Source": "Import"},
    # row10 James White — offer sent, missing amount/date -> EXCEPTION (rule 3)
    {"Property ID": "TVL-0009", "Property Address": "39224 Guardino Dr Apt 208, Fremont, CA 94538",
     "Seller Name": "James White", "Phone": "(209) 221-1240", "Lead Source": "PPL - Property Leads",
     "Visit Date": D(2026,7,14), "Visit Status": "Completed", "Assigned Visitor": "Cesar",
     "Visit Notes": "Offer emailed; James (trustee) + sister Lisa reviewing together. Cherry following up.",
     "Seller Motivation": "Trustee sale; reviewing offer with family",
     "Current Stage": "Offer Sent", "Offer Status": "Sent",
     "Next Action": "Follow up on offer with James/Lisa", "Next Action Due Date": D(2026,7,17),
     "Assigned Owner": "Cherry", "Last Contact Date": D(2026,7,15),
     "Last Contact Result": "Cherry emailing offer breakdown; family reviewing",
     "Blocker": "Family",
     "Created Date": D(2026,7,14), "Last Updated Date": D(2026,7,15), "Updated By": "Import",
     "Source": "Import", "REI Update Required": "Yes"},
    # row11 Mark Lempert — lost (service failure) -> revival candidate
    {"Property ID": "TVL-0010", "Property Address": "1323 Oxford St, Berkeley, CA 94709",
     "Seller Name": "Mark Lempert", "Phone": "(510) 816-1221", "Lead Source": "Direct Mail - Postcard",
     "Visit Date": D(2026,7,13), "Visit Status": "Completed", "Assigned Visitor": "Cesar",
     "Visit Notes": "SERVICE FAILURE: walkthrough done 7/13 but no offer ever sent; seller displeased.",
     "Seller Motivation": "Was open to offer; lost due to our delay/no-offer",
     "Current Stage": "Lost / Closed Out", "Final Disposition": "Lost",
     "Closeout Reason": "Did Not Proceed — service failure (no offer sent)", "Blocker": "Documents",
     "Created Date": D(2026,7,13), "Last Updated Date": D(2026,7,15), "Updated By": "Import", "Source": "Import"},
]

# TEST records (Source=TEST) to exercise the healthy path & every board section.
TEST = [
    {"Property ID": "TEST-01", "Property Address": "100 Test Verbal Ln, Testville, CA 90001",
     "Seller Name": "Val Verbal", "Phone": "(000) 000-0001", "Lead Source": "PPC",
     "REI BlackBook Link": "https://app.reiblackbook.com/lead/test01",
     "Visit Date": D(2026,7,18), "Visit Status": "Completed", "Assigned Visitor": "Juan",
     "Visit Notes": "Great condition; seller keen.", "Seller Motivation": "Relocating for job — motivated",
     "Approved Offer Amount": 640000, "Offer Status": "Accepted", "Offer Sent Date": D(2026,7,19),
     "Current Stage": "Verbal Agreement", "Next Action": "Prepare purchase contract",
     "Next Action Due Date": D(2026,7,23), "Assigned Owner": "Kyle",
     "Last Contact Date": D(2026,7,21), "Last Contact Result": "Seller verbally agreed to $640k",
     "Created Date": D(2026,7,18), "Last Updated Date": D(2026,7,21), "Updated By": "Cherry",
     "Source": "TEST", "REI Update Required": "Yes"},
    {"Property ID": "TEST-02", "Property Address": "200 Test Contract Sent Ave, Testville, CA 90002",
     "Seller Name": "Sam Sent", "Phone": "(000) 000-0002", "Lead Source": "TV",
     "REI BlackBook Link": "https://app.reiblackbook.com/lead/test02",
     "Visit Date": D(2026,7,15), "Visit Status": "Completed", "Assigned Visitor": "Juan",
     "Visit Notes": "Clean title; ready to move.", "Seller Motivation": "Downsizing — motivated",
     "Approved Offer Amount": 720000, "Offer Status": "Accepted", "Offer Sent Date": D(2026,7,16),
     "Current Stage": "Contract Sent", "Contract Sent Date": D(2026,7,20),
     "File Link": "https://drive.google.com/test02-contract",
     "Next Action": "Confirm signature", "Next Action Due Date": D(2026,7,22), "Assigned Owner": "Cherry",
     "Last Contact Date": D(2026,7,21), "Last Contact Result": "Seller reviewing contract",
     "Created Date": D(2026,7,15), "Last Updated Date": D(2026,7,21), "Updated By": "Cherry",
     "Source": "TEST", "REI Update Required": "Yes"},
    {"Property ID": "TEST-03", "Property Address": "300 Test Signed Blvd, Testville, CA 90003",
     "Seller Name": "Sid Signed", "Phone": "(000) 000-0003", "Lead Source": "Direct Mail",
     "REI BlackBook Link": "https://app.reiblackbook.com/lead/test03",
     "Visit Date": D(2026,7,10), "Visit Status": "Completed", "Assigned Visitor": "Juan",
     "Visit Notes": "Signed.", "Seller Motivation": "Estate sale — motivated",
     "Approved Offer Amount": 555000, "Offer Status": "Accepted", "Offer Sent Date": D(2026,7,11),
     "Current Stage": "Contract Signed", "Final Disposition": "Contracted",
     "Contract Sent Date": D(2026,7,14), "Contract Signed Date": D(2026,7,18),
     "Transaction Handoff Status": "Ready for Handoff",
     "Next Action": "Hand off signed contract to JM", "Next Action Due Date": D(2026,7,20),
     "Assigned Owner": "JM", "Last Contact Date": D(2026,7,18), "Last Contact Result": "Contract signed",
     "Created Date": D(2026,7,10), "Last Updated Date": D(2026,7,18), "Updated By": "Cherry",
     "Source": "TEST", "REI Update Required": "Yes"},
    {"Property ID": "TEST-04", "Property Address": "400 Test Needs Review St, Testville, CA 90004",
     "Seller Name": "Nia Needs-Review", "Phone": "(000) 000-0004", "Lead Source": "PPC",
     "REI BlackBook Link": "https://app.reiblackbook.com/lead/test04",
     "Visit Date": D(2026,7,21), "Visit Status": "Completed", "Assigned Visitor": "Juan",
     "Visit Notes": "Visit done; needs offer/pass decision.", "Seller Motivation": "Job relocation",
     "Current Stage": "Visit Completed — Needs Review",
     "Next Action": "Decide: make offer or pass", "Next Action Due Date": D(2026,7,21),
     "Assigned Owner": "Jonathan", "Last Contact Date": D(2026,7,21),
     "Last Contact Result": "Walkthrough complete",
     "Created Date": D(2026,7,21), "Last Updated Date": D(2026,7,21), "Updated By": "Jonathan",
     "Source": "TEST", "REI Update Required": "Yes"},
    {"Property ID": "TEST-05", "Property Address": "500 Test Nurture Way, Testville, CA 90005",
     "Seller Name": "Nora Nurture", "Phone": "(000) 000-0005", "Lead Source": "SEO",
     "REI BlackBook Link": "https://app.reiblackbook.com/lead/test05",
     "Visit Date": D(2026,6,20), "Visit Status": "Completed", "Assigned Visitor": "Juan",
     "Visit Notes": "Not ready yet; call back in ~60 days.", "Seller Motivation": "Will sell later this year",
     "Current Stage": "Long-Term Nurture", "Next Action": "Nurture check-in call",
     "Next Action Due Date": D(2026,9,20), "Assigned Owner": "Cherry",
     "Last Contact Date": D(2026,6,20), "Last Contact Result": "Asked for callback in 60 days",
     "Gift Status": "Recommended", "Gift Recommendation Reason": "Strong rapport; long-term seller",
     "Created Date": D(2026,6,20), "Last Updated Date": D(2026,6,20), "Updated By": "Cherry",
     "Source": "TEST", "REI Update Required": "Yes"},
    {"Property ID": "TEST-06", "Property Address": "600 Test Stalled Ct, Testville, CA 90006",
     "Seller Name": "Stan Stalled", "Phone": "(000) 000-0006", "Lead Source": "PPC",
     "REI BlackBook Link": "https://app.reiblackbook.com/lead/test06",
     "Visit Date": D(2026,7,1), "Visit Status": "Completed", "Assigned Visitor": "Juan",
     "Visit Notes": "Offer sent; seller gone quiet.", "Seller Motivation": "Was motivated; now unresponsive",
     "Approved Offer Amount": 500000, "Offer Status": "Sent", "Offer Sent Date": D(2026,7,2),
     "Current Stage": "Offer Sent", "Next Action": "Re-attempt contact", "Next Action Due Date": D(2026,7,8),
     "Assigned Owner": "Juan", "Blocker": "Seller Unresponsive",
     "Last Contact Date": D(2026,7,2), "Last Contact Result": "Offer sent; no reply since",
     "Created Date": D(2026,7,1), "Last Updated Date": D(2026,7,2), "Updated By": "Juan",
     "Source": "TEST", "REI Update Required": "Yes"},
    {"Property ID": "TEST-07", "Property Address": "700 Test Negotiation Rd, Testville, CA 90007",
     "Seller Name": "Neil Negotiate", "Phone": "(000) 000-0007", "Lead Source": "TV",
     "REI BlackBook Link": "https://app.reiblackbook.com/lead/test07",
     "Visit Date": D(2026,7,16), "Visit Status": "Completed", "Assigned Visitor": "Juan",
     "Visit Notes": "Seller countered.", "Seller Motivation": "Motivated but wants more money",
     "Approved Offer Amount": 610000, "Offer Status": "Countered", "Offer Sent Date": D(2026,7,17),
     "Counteroffer Amount": 660000,
     "Current Stage": "Active Negotiation", "Next Action": "Cherry/Juan decide counter response",
     "Next Action Due Date": D(2026,7,22), "Assigned Owner": "Cherry", "Blocker": "Price",
     "Last Contact Date": D(2026,7,21), "Last Contact Result": "Seller countered at $660k",
     "Created Date": D(2026,7,16), "Last Updated Date": D(2026,7,21), "Updated By": "Cherry",
     "Source": "TEST", "REI Update Required": "Yes"},
    # duplicate-address test: same normalized address as TEST-07, still active -> DUPLICATE flag on both
    {"Property ID": "TEST-08", "Property Address": "700 Test Negotiation Rd., Testville, CA 90007",
     "Seller Name": "Duplicate Entry", "Phone": "(000) 000-0008", "Lead Source": "TV",
     "REI BlackBook Link": "https://app.reiblackbook.com/lead/test08",
     "Visit Date": D(2026,7,16), "Visit Status": "Scheduled", "Assigned Visitor": "Juan",
     "Current Stage": "Visit Scheduled", "Next Action": "Verify duplicate", "Next Action Due Date": D(2026,7,23),
     "Assigned Owner": "Cherry", "Created Date": D(2026,7,16), "Last Updated Date": D(2026,7,16),
     "Updated By": "Cherry", "Source": "TEST", "REI Update Required": "Yes"},
    # dormant lost lead -> exercises Revival Opportunities (Lost + >=45 days since activity)
    {"Property ID": "TEST-09", "Property Address": "900 Test Revival Dr, Testville, CA 90009",
     "Seller Name": "Rita Revival", "Phone": "(000) 000-0009", "Lead Source": "Direct Mail",
     "REI BlackBook Link": "https://app.reiblackbook.com/lead/test09",
     "Visit Date": D(2026,4,1), "Visit Status": "Completed", "Assigned Visitor": "Juan",
     "Visit Notes": "Passed in spring; seller wanted more than numbers allowed.",
     "Seller Motivation": "Was motivated; timing/price gap at the time",
     "Current Stage": "Lost / Closed Out", "Final Disposition": "Lost",
     "Closeout Reason": "Seller Rejected Offer — price gap (revisit later)",
     "Last Contact Date": D(2026,4,3), "Last Contact Result": "Seller declined; open to future contact",
     "Created Date": D(2026,4,1), "Last Updated Date": D(2026,4,3), "Updated By": "Cherry",
     "Source": "TEST"},
]

ALL_RECORDS = PILOT + TEST


# ---------------------------------------------------------------------------
# 4. BUILD WORKBOOK
# ---------------------------------------------------------------------------
def build():
    wb = openpyxl.Workbook()
    ARIAL = "Arial"

    # ---- Data sheet ----
    data = wb.active
    data.title = "Data"
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # group banner row (row 1) + header row (row 2)
    # We'll put headers on row 1 to keep it simple & Apps-Script-friendly (data starts row 2).
    for i, (group, header, dd, fx) in enumerate(COLUMNS):
        col = i + 1
        c = data.cell(1, col, header)
        c.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=_hdrcolor(group))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        data.column_dimensions[get_column_letter(col)].width = _width(header)
    data.freeze_panes = "A2"
    data.row_dimensions[1].height = 34

    # write records
    firstrow, lastrow = 2, 1 + len(ALL_RECORDS)
    for ridx, rec in enumerate(ALL_RECORDS):
        row = firstrow + ridx
        for i, (group, header, dd, fx) in enumerate(COLUMNS):
            col = i + 1
            cell = data.cell(row, col)
            cell.font = Font(name=ARIAL, size=10)
            cell.border = border
            if fx:
                cell.value = render_formula(fx, row)
            elif header in rec:
                cell.value = rec[header]
            # date formatting
            if isinstance(rec.get(header), datetime.datetime):
                cell.number_format = "yyyy-mm-dd"
            if header in ("Asking Price", "Price Expectation", "Approved Offer Amount", "Counteroffer Amount"):
                cell.number_format = '$#,##0'

    # extend formula columns down to row 60 (blank rows guard themselves)
    for row in range(lastrow + 1, 61):
        for i, (group, header, dd, fx) in enumerate(COLUMNS):
            if fx:
                cell = data.cell(row, i + 1)
                cell.value = render_formula(fx, row)
                cell.font = Font(name=ARIAL, size=10)

    # data validations (dropdowns) applied to rows 2..1000
    for i, (group, header, dd, fx) in enumerate(COLUMNS):
        if dd and dd in DROPDOWNS:
            letter = get_column_letter(i + 1)
            vals = ",".join(DROPDOWNS[dd])
            dv = DataValidation(type="list", formula1=f'"{vals}"', allow_blank=True, showDropDown=False)
            dv.error = "Pick a value from the list."
            dv.prompt = f"Allowed: {vals[:200]}"
            data.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}1000")

    # conditional formatting on Data: overdue red, stalled orange, signed green
    from openpyxl.formatting.rule import FormulaRule, CellIsRule
    red = PatternFill("solid", fgColor="F4CCCC"); redf = Font(color="990000")
    orange = PatternFill("solid", fgColor="FCE5CD"); orangef = Font(color="B45F06")
    green = PatternFill("solid", fgColor="D9EAD3"); greenf = Font(color="38761D")
    dq_col = COL_LET["Data Quality Status"]
    ov_col = COL_LET["Days Overdue"]
    st_col = COL_LET["Stalled Status"]
    cs_col = COL_LET["Current Stage"]
    rng = f"A2:{get_column_letter(len(COLUMNS))}60"
    data.conditional_formatting.add(rng,
        FormulaRule(formula=[f'${dq_col}2="Exception"'], fill=red, font=redf, stopIfTrue=False))
    data.conditional_formatting.add(rng,
        FormulaRule(formula=[f'AND(${ov_col}2<>"",${ov_col}2>0)'], fill=red, stopIfTrue=False))
    data.conditional_formatting.add(rng,
        FormulaRule(formula=[f'${st_col}2="Yes"'], fill=orange, font=orangef, stopIfTrue=False))
    data.conditional_formatting.add(rng,
        FormulaRule(formula=[f'${dq_col}2="Incomplete"'], fill=orange, stopIfTrue=False))
    data.conditional_formatting.add(rng,
        FormulaRule(formula=[f'${cs_col}2="Contract Signed"'], fill=green, font=greenf, stopIfTrue=False))

    _build_readme_sheet(wb, ARIAL, firstrow, lastrow)
    _build_dropdowns_sheet(wb, ARIAL)
    _build_board_sheet(wb, ARIAL, firstrow)
    _build_exceptions_sheet(wb, ARIAL)
    _build_migration_log(wb, ARIAL)

    out = "/home/user/Property-Visit-Tracking/build/Twin_Visit_Logger_DEV_reference.xlsx"
    wb.save(out)
    print("Saved", out)
    print("Columns:", len(COLUMNS), "| Records:", len(ALL_RECORDS),
          f"(pilot {len(PILOT)} + test {len(TEST)}) | data rows {firstrow}-{lastrow}")


def _hdrcolor(group):
    base = {"Property":"2E5A88","Visit":"548235","Seller":"BF9000","Offer":"C55A11",
            "Follow-up":"2E75B6","Relationship":"A64D79","Closeout":"674EA7",
            "Computed":"7F7F7F","System":"595959"}
    return base[group]


def _width(h):
    if h in ("Visit Notes","Seller Concerns","Seller Motivation","Last Contact Result",
             "Next Action","Closeout Reason","Missing Required Fields","Exception Reason",
             "Gift Recommendation Reason","Property Address","REI BlackBook Link","File Link",
             "Photos Link","Video Link"):
        return 34
    return 16


def _build_readme_sheet(wb, ARIAL, firstrow, lastrow):
    ws = wb.create_sheet("READ ME", 0)
    ws.sheet_properties.tabColor = "38761D"
    lines = [
        ("Twin Visit Logger — Development Copy (Reference Workbook)", True, 14),
        ("", False, 10),
        ("This reference workbook demonstrates the Phase 2 upgraded structure.", False, 10),
        ("The LIVE system is the Google Sheets dev copy; run apps-script/Setup.gs there to build it natively.", False, 10),
        ("", False, 10),
        ("Sheets:", True, 11),
        ("  • Data — main pipeline (one row per property). Data starts on row 2.", False, 10),
        ("  • Cherry Opportunity Board — the 30-second action view (10 sections).", False, 10),
        ("  • Dropdowns — reference lists behind every data-validation field.", False, 10),
        ("  • Exception Queue — records needing review (incomplete / rule failures / duplicates).", False, 10),
        ("  • Migration Log — how legacy fields map to the new structure.", False, 10),
        ("", False, 10),
        (f"Pilot + test records: rows {firstrow}–{lastrow} on Data.", False, 10),
        ("Original workbook was NOT modified. This is a separate development copy.", True, 10),
        ("", False, 10),
        ("Color key:  Red = error/overdue   Orange = warning/incomplete/stalled   Green = signed/complete", False, 10),
    ]
    for i,(t,b,s) in enumerate(lines, start=1):
        c = ws.cell(i,1,t); c.font = Font(name=ARIAL, bold=b, size=s)
    ws.column_dimensions["A"].width = 110


def _build_dropdowns_sheet(wb, ARIAL):
    ws = wb.create_sheet("Dropdowns")
    col = 1
    for key, vals in DROPDOWNS.items():
        h = ws.cell(1, col, key)
        h.font = Font(name=ARIAL, bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="595959")
        for r, v in enumerate(vals, start=2):
            ws.cell(r, col, v).font = Font(name=ARIAL, size=10)
        ws.column_dimensions[get_column_letter(col)].width = 26
        col += 1


def _build_exceptions_sheet(wb, ARIAL):
    ws = wb.create_sheet("Exception Queue")
    ws.sheet_properties.tabColor = "990000"
    title = ws.cell(1,1,"Exception Queue — records that are Incomplete, fail a cross-field rule, or duplicate")
    title.font = Font(name=ARIAL, bold=True, size=12)
    hdrs = ["Property ID","Property Address","Seller Name","Current Stage","Assigned Owner",
            "Data Quality Status","Missing Required Fields","Exception Reason"]
    for c,h in enumerate(hdrs, start=1):
        cell = ws.cell(3,c,h); cell.font = Font(name=ARIAL,bold=True,color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="990000")
        ws.column_dimensions[get_column_letter(c)].width = 30 if h in ("Missing Required Fields","Exception Reason","Property Address") else 18
    # Google Sheets QUERY (native). Documented; not evaluated by LibreOffice.
    q = ('=IFERROR(QUERY(Data!A2:BZ1000,'
         '"select {A},{B},{D},{AQ},{AH} , {BE},{AW},{BF} '
         'where B is not null and (BE=\'Incomplete\' or BE=\'Exception\') '
         'label ...",0),"")')
    note = ws.cell(4,1,
        "In the live Google Sheet this list is generated by a QUERY over Data "
        "(Data Quality Status = Incomplete or Exception). See Data-Dictionary.md.")
    note.font = Font(name=ARIAL, italic=True, size=9, color="666666")


def _build_migration_log(wb, ARIAL):
    ws = wb.create_sheet("Migration Log")
    ws.sheet_properties.tabColor = "BF9000"
    rows = [
        ["Legacy field","New field(s)","Mapping rule","Confidence / notes"],
        ["Address","Property Address (+ Normalized Address)","Copied verbatim; Normalized Address computed","High"],
        ["Name","Seller Name","Copied verbatim","High"],
        ["Phone","Phone","Copied verbatim","High"],
        ["Lead Source","Lead Source","Values already match new dropdown","High"],
        ["Appointment date / col A","Visit Date","Copied","High"],
        ["Inspection Status","Visit Status","Inspected→Completed; Pending Inspection→Scheduled; Cancelled→Canceled; 'Skipped - offer made'→Completed","High"],
        ["Inspector","Assigned Visitor","Juan Diaz→Juan; Cesar→Cesar (kept); others kept as-is in visitor list","Medium"],
        ["Closer / Agent","Assigned Owner","Cherry→Cherry; Juan Diaz→Juan; blank/other→left blank → Incomplete → Exception Queue","Low where blank"],
        ["Deal Stage + Deal Status","Current Stage (+ Final Disposition)","See stage-mapping table in Data-Dictionary.md","Medium — uncertain rows sent to Exception Queue"],
        ["Status Update (prose)","Last Contact Result / Next Action / Visit Notes","Clear 'next step' text mapped to Next Action; full text preserved in Visit Notes; ambiguous → Exception","Low — not guessed"],
        ["Notes","Visit Notes / Seller Motivation","Notes→Visit Notes; motivation only where explicit","Medium"],
        ["Golden Needle (unused, all FALSE)","(dropped)","Audit flagged as unused; not migrated","n/a"],
        ["Contract (dropdown)","Final Disposition / Transaction Handoff Status","Under Contract→(context); Acquired→Contracted; Cancelled Contract→Lost","Medium"],
    ]
    for r, row in enumerate(rows, start=1):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(r,c,v)
            cell.font = Font(name=ARIAL, bold=(r==1), size=10, color=("FFFFFF" if r==1 else "000000"))
            if r==1: cell.fill = PatternFill("solid", fgColor="BF9000")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for c,w in enumerate([22,34,52,34], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_board_sheet(wb, ARIAL, firstrow):
    """Cherry Opportunity Board — section headers + Google QUERY formulas (native).
    LibreOffice cannot evaluate QUERY, so these are written as strings and NOT
    locally recalculated; they are correct for Google Sheets. Board display columns:
    Address, Seller, Current Stage, Next Action, Owner, Due Date, Days Overdue,
    Blocker, Last Contact Result, REI Link.
    """
    ws = wb.create_sheet("Cherry Opportunity Board", 1)
    ws.sheet_properties.tabColor = "2E75B6"
    C = COL_LET
    disp = ["Property Address","Seller Name","Current Stage","Next Action","Assigned Owner",
            "Next Action Due Date","Days Overdue","Blocker","Last Contact Result","REI BlackBook Link"]
    # QUERY select column letters
    sel = ",".join(C[h] for h in disp)

    title = ws.cell(1,1,"CHERRY OPPORTUNITY BOARD — actionable opportunities only")
    title.font = Font(name=ARIAL, bold=True, size=14, color="1F4E79")
    sub = ws.cell(2,1,"Updated live from Data. Sections sorted by contract-likelihood, overdue, nearest due date, recent engagement.")
    sub.font = Font(name=ARIAL, italic=True, size=9, color="666666")

    stage = C["Current Stage"]; due = C["Next Action Due Date"]; ov = C["Days Overdue"]
    stall = C["Stalled Status"]; prio = C["Opportunity Priority"]; dq = C["Data Quality Status"]
    gift = C["Gift Status"]; handoff = C["Transaction Handoff Status"]; owner = C["Assigned Owner"]
    disp_val = C["Final Disposition"]

    # section (title, where-clause, order-by)
    sections = [
        ("1. Contracts Possible This Week",
         f"({stage}='Verbal Agreement' or {stage}='Contract Sent' or {stage}='Active Negotiation')",
         f"{prio} desc, {due}"),
        ("2. Visited — No Offer Decision",
         f"{stage}='Visit Completed — Needs Review'", f"{ov} desc, {due}"),
        ("3. Offer Sent — Follow-Up Due",
         f"{stage}='Offer Sent'", f"{ov} desc, {due}"),
        ("4. Stalled Deals",
         f"{stall}='Yes'", f"{ov} desc, {prio} desc"),
        ("5. Overdue Tasks",
         f"{ov}>0", f"{ov} desc"),
        ("6. Negotiation Decisions",
         f"{stage}='Active Negotiation'", f"{prio} desc, {due}"),
        ("7. Contract Handoffs",
         f"({stage}='Contract Signed' and {handoff}<>'JM Confirmed')", f"{due}"),
        ("8. Gift Review",
         f"{gift}='Recommended'", f"{due}"),
        ("9. Revival Opportunities",
         f"({disp_val}='Lost' and {C['Days Since Last Activity']}>=45)", f"{C['Days Since Last Activity']} desc"),
        ("10. Exceptions Requiring Review",
         f"({dq}='Exception' or {dq}='Incomplete')", f"{stage}"),
    ]

    row = 4
    hdr_titles = ["Address","Seller","Stage","Next Action","Owner","Due","Days Overdue","Blocker","Last Contact Result","REI Link"]
    for sec_title, where, order in sections:
        sc = ws.cell(row,1,sec_title)
        sc.font = Font(name=ARIAL, bold=True, size=12, color="FFFFFF")
        sc.fill = PatternFill("solid", fgColor="2E75B6")
        for cc in range(2, 11):
            ws.cell(row,cc).fill = PatternFill("solid", fgColor="2E75B6")
        row += 1
        for c,h in enumerate(hdr_titles, start=1):
            hc = ws.cell(row,c,h)
            hc.font = Font(name=ARIAL, bold=True, size=9)
            hc.fill = PatternFill("solid", fgColor="DDEBF7")
        row += 1
        q = (f'=IFERROR(QUERY(Data!A{firstrow}:BZ1000,'
             f'"select {sel} where {C["Property Address"]} is not null and {where} '
             f'order by {order} limit 50",0),"— none —")')
        qc = ws.cell(row,1,q)
        qc.font = Font(name=ARIAL, size=10)
        row += 8   # leave space for spilled results
    ws.column_dimensions["A"].width = 32
    for cl in ["B","C","D","E","F","G","H","I","J"]:
        ws.column_dimensions[cl].width = 18
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["I"].width = 34
    ws.column_dimensions["J"].width = 30

    # filter helper note
    note_row = row + 1
    n = ws.cell(note_row,1,
        "QUICK FILTERS (live sheet): use the toolbar Filter Views 'My Tasks', 'Due Today', 'Overdue', "
        "'Stalled', 'Needs Offer Decision', 'Offer Follow-Up', 'Negotiation Decision', "
        "'Contracts Possible This Week', 'Gift Review', 'Exceptions' — created by Setup.gs.")
    n.font = Font(name=ARIAL, italic=True, size=9, color="666666")


if __name__ == "__main__":
    build()
