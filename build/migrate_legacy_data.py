"""
Convert the live "Property Visit Tracking" workbook's Data tab into the Twin Visit Logger
tracker layout, so every historical lead can be imported into the DEV sheet in one paste.

    python build/migrate_legacy_data.py SOURCE.xlsx --out build/legacy-import.csv

What it does NOT do: it never writes to Google. It produces a CSV whose header row matches the
tracker's own headers exactly; the workbook's "Import legacy rows" menu item reads that CSV by
header name, so column order in the paste does not matter and the 9 computed formula columns are
left untouched.

Every mapping decision below is stated explicitly. Nothing is invented: where the legacy sheet has
no answer, the cell is left blank so the record surfaces in the dashboard's
"Unrouted - Needs Attention" section for a human to triage.
"""
import argparse
import csv
import datetime
import re
import sys

import openpyxl

# ---------------------------------------------------------------------------
# Tracker layout. Keep IN SYNC with apps-script/Config.gs HEADERS.
# ---------------------------------------------------------------------------
HEADERS = [
    'Property ID', 'Property Address', 'Normalized Address', 'Seller Name', 'Phone', 'Email',
    'Lead Source', 'REI BlackBook Link',
    'Visit Date', 'Visit Time', 'Visit Status', 'Assigned Visitor', 'Visit Notes',
    'Property Condition', 'Occupancy Status', 'Photos Link', 'Video Link', 'File Link',
    'Seller Motivation', 'Seller Timeline', 'Asking Price', 'Price Expectation', 'Seller Concerns',
    'Approved Offer Amount', 'Offer Status', 'Offer Prepared Date', 'Offer Sent Date',
    'Offer Received Confirmation', 'Counteroffer Amount',
    'Last Contact Date', 'Last Contact Result', 'Next Action', 'Next Action Due Date',
    'Assigned Owner', 'Blocker', 'Days Since Last Activity', 'Days Overdue', 'Stalled Status',
    'Gift Status', 'Gift Recommendation Reason', 'Gift Approval Owner', 'Gift Sent Date',
    'Current Stage', 'Final Disposition', 'Closeout Reason', 'Contract Sent Date',
    'Contract Signed Date', 'Transaction Handoff Status',
    'Missing Required Fields', 'Duplicate Address Flag', 'Opportunity Priority',
    'Created Date', 'Last Updated Date', 'Updated By', 'Source', 'Data Quality Status',
    'Exception Reason', 'REI Update Required', 'REI Update Completed',
    'Gift Approved By', 'Gift Approval Date', 'Offer Promised Date', 'Seller Floor', 'Our Max',
    # Appended for the legacy migration - fields the old workbook tracked that had no home here.
    'City', 'Deal Stage', 'Deal Status', 'Contract Status', 'Closer', 'Golden Needle',
    'Market Status Update',
    # Shared with the office-PC automation, which writes the ID of the calendar event it created.
    # The migration never fills it: a legacy row has no event, and inventing one would make the
    # Apps Script skip creating a real event for any visit that is still to happen.
    'Calendar Event ID',
]

# The sheet owns these - they are formulas. Never write them.
COMPUTED = {
    'Normalized Address', 'Days Since Last Activity', 'Days Overdue', 'Stalled Status',
    'Missing Required Fields', 'Duplicate Address Flag', 'Opportunity Priority',
    'Data Quality Status', 'Exception Reason',
}

# Legacy Data tab column order (1-based), as found in the live workbook.
LEGACY = {
    'created': 1, 'name': 2, 'phone': 3, 'address': 4, 'city': 5, 'inspection': 6,
    'source': 7, 'contract': 8, 'stage': 9, 'status': 10, 'appointment': 11,
    'inspector': 12, 'closer': 13, 'golden': 14, 'agent': 15, 'notes': 16,
    'market': 17, 'lastupdate': 18,
}

# ---------------------------------------------------------------------------
# Value maps
# ---------------------------------------------------------------------------

# Legacy "Inspection Status" -> tracker "Visit Status".
VISIT_STATUS = {
    'inspected': 'Completed',
    'cancelled': 'Canceled',          # tracker spells it with one L
    'canceled': 'Canceled',
    'pending inspection': 'Scheduled',
    'skipped - offer made': 'Skipped — Offer Made',
}

# Legacy "Deal Stage" -> canonical four, per the workbook's own
# "Ref (Deals) - Tags definition" tab. The live data carries spelling drift.
DEAL_STAGE = {
    'active': 'Active',
    'on hold': 'On Hold',
    'won (closed)': 'Won',
    'won': 'Won',
    'lost': 'Lost',
}

# Legacy "Deal Status" -> the canonical spelling on the Tags definition tab.
DEAL_STATUS = {
    'lead received': 'Lead Received',
    'appointment scheduled': 'Appointment Scheduled',
    'pending reschedule': 'Pending Reschedule',
    'under review': 'Under Review',
    'offer made': 'Offer Made',
    'under contract': 'Under Contract',
    'on hold - follow up scheduled': 'On Hold - Follow Up Scheduled',
    'on hold - nurture': 'On Hold - Nurture',
    'on hold - awaiting seller': 'On Hold - Awaiting Seller',
    'on hold - probate/legal': 'On Hold - Probate/Legal',
    'on hold - seller timeline': 'On Hold - Seller Timeline',
    'acquired': 'Acquired',
    'acquired - in rehab': 'Acquired - In Rehab',
    'acquired - listed': 'Acquired - Listed',
    'acquired - sold': 'Acquired - Sold',
    'wholesale - buyer assigned': 'Wholesale - Buyer Assigned',
    'wholesale - deal closed': 'Wholesale - Deal Closed',
    'not qualified': 'Not Qualified',
    "we're passing": "We're Passing",
    'contract cancelled': 'Contract Cancelled',
    'seller rejected offer': 'Seller Rejected Offer',
    'did not proceed': 'Did Not Proceed',
    'sold to competitor': 'Sold to Competitor',
    'sold with realtor': 'Sold with Realtor',
    'referred to realtor': 'Referred to Realtor',
    'already listed': 'Already listed',
    'sold (unknown buyer)': 'Sold (unknown buyer)',
}

# The legacy "Agent" column is free text and a few cells carry an explanation rather than a name
# ("Matt-since it was Juan"). Assigned Owner is a validated dropdown, so the name is extracted and
# the explanation is preserved in Visit Notes instead of being dropped or breaking validation.
AGENT_NAMES = ['Matt/Arly', 'Matt/Juan', 'Cherry/Matt', 'Jonathan', 'Danica', 'Darius', 'Cherry',
               'Team', 'Arly', 'Matt', 'Kyle', 'Juan']


def split_agent(raw):
    """Return (assigned_owner, leftover_note). Either may be ''."""
    if not raw:
        return '', ''
    for name in AGENT_NAMES:                      # longest/compound names first
        if raw.lower().startswith(name.lower()):
            rest = raw[len(name):].strip(' -–—:;,')
            return name, rest
    return '', raw                                # unrecognised: keep it as a note, not as an owner


# Deal Status -> tracker Current Stage, for rows whose Deal Stage is "Active".
ACTIVE_STATUS_STAGE = {
    'Under Contract': 'Contract Signed',
    'Offer Made': 'Offer Sent',
    'Under Review': 'Offer Preparation',
    'Lead Received': 'Visit Scheduled',
    'Appointment Scheduled': 'Visit Scheduled',
    'Pending Reschedule': 'Visit Scheduled',
    'Seller Rejected Offer': 'Lost / Closed Out',
    'Did Not Proceed': 'Lost / Closed Out',
}


def text(value):
    if value is None:
        return ''
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r'[ \t]+', ' ', str(value)).strip()


def date_only(value):
    """Dates go in as YYYY-MM-DD so Sheets parses them as real dates, not text."""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.date().isoformat() if isinstance(value, datetime.datetime) else value.isoformat()
    return ''


def current_stage(stage, status, inspection, contract):
    """
    Decide the tracker stage. Order matters: a contract outranks a deal stage, which outranks the
    inspection result. Returns '' when the legacy data does not actually say - those rows land in
    "Unrouted - Needs Attention" instead of being given a stage nobody chose.
    """
    if contract == 'Acquired':
        return 'Contract Signed'
    if contract == 'Under Contract':
        return 'Contract Signed'
    if contract == 'Cancelled Contract':
        return 'Lost / Closed Out'

    if stage == 'Lost':
        return 'Lost / Closed Out'
    if stage == 'Won':
        return 'Contract Signed'
    if stage == 'On Hold':
        return 'Long-Term Nurture'
    if stage == 'Active':
        if status.startswith('On Hold'):
            return 'Long-Term Nurture'
        if status.startswith('Acquired') or status.startswith('Wholesale'):
            return 'Contract Signed'
        mapped = ACTIVE_STATUS_STAGE.get(status)
        if mapped:
            return mapped
        if inspection == 'Pending Inspection':
            return 'Visit Scheduled'
        if inspection == 'Inspected':
            return 'Visit Completed — Needs Review'
        return ''

    # No deal stage recorded. Only the unambiguous inspection results imply a stage.
    if inspection == 'Inspected':
        return 'Visit Completed — Needs Review'
    if inspection == 'Pending Inspection':
        return 'Visit Scheduled'
    return ''   # includes "Cancelled with no deal stage" - genuinely needs a human


def final_disposition(stage, contract):
    if contract == 'Acquired' or stage == 'Won':
        return 'Contracted'
    if stage == 'Lost' or contract == 'Cancelled Contract':
        return 'Lost'
    if stage == 'On Hold':
        return 'Long-Term Nurture'
    return ''


def build_rows(worksheet, start_id):
    out, warnings = [], []
    seq = start_id
    for r in range(2, worksheet.max_row + 1):
        cell = lambda key: worksheet.cell(r, LEGACY[key]).value  # noqa: E731

        name = text(cell('name'))
        address = text(cell('address'))
        if not name and not address:
            continue  # genuinely empty spreadsheet row

        inspection = text(cell('inspection'))
        stage = DEAL_STAGE.get(text(cell('stage')).lower(), '')
        raw_status = text(cell('status'))
        status = DEAL_STATUS.get(raw_status.lower(), raw_status)
        contract = text(cell('contract'))
        visit_status = VISIT_STATUS.get(inspection.lower(), '')
        stage_out = current_stage(stage, status, inspection, contract)

        if raw_status and raw_status.lower() not in DEAL_STATUS:
            warnings.append(f'row {r}: Deal Status "{raw_status}" is not in the Tags definition tab')
        if inspection and not visit_status:
            warnings.append(f'row {r}: Inspection Status "{inspection}" has no Visit Status mapping')
        if not address:
            warnings.append(f'row {r}: no property address ("{name}") - it will import but cannot route')

        owner, owner_note = split_agent(text(cell('agent')))
        if owner_note:
            warnings.append(f'row {r}: Agent "{text(cell("agent"))}" -> owner "{owner or "(none)"}", '
                            f'rest kept in Visit Notes')
        notes = text(cell('notes'))
        if owner_note:
            notes = (notes + ' | ' if notes else '') + 'Agent note: ' + owner_note

        record = {
            'Property ID': f'TVL-{seq:04d}',
            'Property Address': address,
            'Seller Name': name,
            'Phone': text(cell('phone')),
            'Lead Source': text(cell('source')),
            'Visit Date': date_only(cell('appointment')),
            'Visit Status': visit_status,
            'Assigned Visitor': text(cell('inspector')),
            'Visit Notes': notes,
            'Last Contact Date': date_only(cell('lastupdate')),
            'Assigned Owner': owner,
            'Current Stage': stage_out,
            'Final Disposition': final_disposition(stage, contract),
            'Closeout Reason': status if stage == 'Lost' else '',
            'Created Date': date_only(cell('created')),
            'Last Updated Date': date_only(cell('lastupdate')) or date_only(cell('created')),
            'Updated By': 'Import',
            'Source': 'Import',
            'City': text(cell('city')),
            'Deal Stage': stage,
            'Deal Status': status,
            'Contract Status': contract,
            'Closer': text(cell('closer')),
            'Golden Needle': 'Yes' if text(cell('golden')).lower() == 'true' else '',
            'Market Status Update': text(cell('market')),
        }
        out.append(record)
        seq += 1
    return out, warnings


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('source', help='the live Property Visit Tracking .xlsx')
    parser.add_argument('--out', default='build/legacy-import.csv')
    parser.add_argument('--sheet', default='Data')
    parser.add_argument('--start-id', type=int, default=1001,
                        help='first Property ID number; 1001 keeps imports clear of the TVL-00xx pilot rows')
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.source, data_only=True)
    if args.sheet not in workbook.sheetnames:
        sys.exit(f'No "{args.sheet}" tab in {args.source}. Tabs: {workbook.sheetnames}')

    rows, warnings = build_rows(workbook[args.sheet], args.start_id)
    writable = [h for h in HEADERS if h not in COMPUTED]

    with open(args.out, 'w', newline='', encoding='utf-8') as handle:
        writer = csv.DictWriter(handle, fieldnames=writable, extrasaction='ignore')
        writer.writeheader()
        for record in rows:
            writer.writerow({h: record.get(h, '') for h in writable})

    staged = sum(1 for r in rows if r['Current Stage'])
    print(f'{len(rows)} record(s) -> {args.out}')
    print(f'  {staged} carry a stage; {len(rows) - staged} have none and will land in '
          f'"Unrouted - Needs Attention" for triage')
    print(f'  {sum(1 for r in rows if not r["Property Address"])} have no address')
    print(f'  {sum(1 for r in rows if r["Visit Date"])} have an appointment date')
    if warnings:
        print(f'\n{len(warnings)} warning(s):')
        for line in warnings[:25]:
            print('  ' + line)
        if len(warnings) > 25:
            print(f'  ... and {len(warnings) - 25} more')


if __name__ == '__main__':
    main()
